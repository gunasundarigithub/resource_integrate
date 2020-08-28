import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, colors
import os
import sys
sys.path.append('C:\\ctpt\\app-shift-mgnt\\backend')
sys.path.append('C:\\ctpt\\app-shift-mgnt')
from format_excel import formatExcel
from util import util

const = util.excelConstants()
cfg = util.get_conf()

class ACCShiftPlan():
    def __init__(self, asso_list, *args):
        proj_dir = os.path.dirname(os.path.abspath(__file__))
        #self.excel_file = os.path.join(proj_dir, 'CTPT ACC SHIFT PLAN.xlsx')
        self.log = util.get_logger_obj()
        self.assoicates = asso_list
        #self.accPlan = asso_acc_list
        #self.offPlan = asso_off_list
        #self.assos_roaster_plan = assos_roaster_plan    # dict with key ('asso_plan') and values (dates in tuples)
        self.team = args[0]            
        self.month = args[1]
        self.year = args[2]
        self.shore = args[3]
        self.excel_file = 'CTPT ACC SHIFT PLAN.xlsx'
        self.sheetName = '-'.join([self.team, self.year])
        self.format_excel_inst = formatExcel(self.excel_file, self.sheetName)

    """
    Save the excel dataframe to excel workbook.
    """
    def __save__(self, wb_inst):
        wb_inst.save(self.excel_file)

    """
    Fetch the existing dataframe contents fron the excel workbook.
    """
    def __existing_dataframe__(self):
        self.ex_df = pd.read_excel(self.excel_file, sheet_name=self.sheetName)
        return self.ex_df
    
    """
    Check if excel sheet exists, if not, creates a new excel sheet
    """
    def check_excel_exists(self):
        if not os.path.exists(self.excel_file):
            # If file not exists, create excel using pandas.
            self.log.info('Excel file not existing!!, Creating....')
            self.create_excel()
            self.log.info('File created....')
        else:
            # Just make sure excel has sheets (that you defined..)
            wb_existing = Workbook()
            self.log.info(f'Listing out sheets present in the excel: {self.excel_file}')
            for sheet in wb_existing:
                self.log.debug(sheet.title)
        return True
    
    """
    Create Excel if not exists.
    """
    def create_excel(self):
        initial_excel = Workbook()        
        if self.sheetName:
            self.log.debug(f'Adding user defined sheet name: {str(self.sheetName)}')
            initial_excel.create_sheet(self.sheetName, 0)    # Insert at first position.
            self.log.debug('Remove default sheet name..')
            initial_excel.remove(initial_excel['Sheet'])
        initial_excel.save(self.excel_file)

    """
    Function to add overall shift plan headers for a month.
    """
    def add_headers_to_sheet(self):
        self.log.debug(f'MONTH: {self.month}')
        print(util.generate_month_days(self.month))
        if util.generate_month_days(self.month):
            days_list = util.generate_month_days(self.month)
            self.log.debug(f'days in month: {self.month} --> {days_list[0]}')
            first_headers = ['Month', 'Shore', 'Day Number ->']
            first_headers.extend(days_list[0])
            first_headers.append('SHIFT DAYS STATS')
            pre_req_excel = pd.ExcelWriter(self.excel_file, engine='openpyxl')
            # We create first header here for the excel. (Month, Shore, Day Number -->, series of day no.s)
            # We create second header here for the excel. (Team Member, series of Week days)
            header_df1 = self.create_dataframe_for_headers(first_headers, pre_req_excel)
            second_header = util.generate_weekdays_for_month(self.year, self.month, len(days_list[0]))
            header_df2 = self.create_dataframe_for_headers(second_header, pre_req_excel)

            # Create a Pandas Excel writer using XlsxWriter as the engine.
            excel_writer = pd.ExcelWriter(self.excel_file, engine='openpyxl')
            # OPen existng workbook
            excel_writer.book = load_workbook(self.excel_file)
            # copy the existing contents in sheet
            excel_writer.sheets = dict((ex_content.title, ex_content) for ex_content in excel_writer.book.worksheets)
            # Read the excel file
            existing_df = self.__existing_dataframe__()
