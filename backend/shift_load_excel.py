"""
Copy Rights, All Rights Reserved 2020
File name: shift_load_excel.py

Description: 
Script collects the parsed user inputs (shift plan data from UI) in the form of python dict object
and appends the contents to the excel sheet.

Code Changes:
Initial: Sabarish AC
Modified by: Sayan H
"""

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, colors
import os
import sys
sys.path.append('C:\\Sabs Learning\\Python Learning\\resource_integrate\\backend')
sys.path.append('C:\\Sabs Learning\\Python Learning\\resource_integrate')
from format_excel import formatExcel
from util import util

const = util.excelConstants()
cfg = util.get_conf()

class ACCShiftPlan():
    def __init__(self, asso_list, *args):
        proj_dir = os.path.dirname(os.path.abspath(__file__))
        #self.excel_file = os.path.join(proj_dir, 'CTPT ACC SHIFT PLAN.xlsx')
        self.log = util.get_logger_obj()
        self.assoicates = asso_list
        #self.accPlan = asso_acc_list
        #self.offPlan = asso_off_list
        #self.assos_roaster_plan = assos_roaster_plan    # dict with key ('asso_plan') and values (dates in tuples)
        self.team = args[0]            
        self.month = args[1]
        self.year = args[2]
        self.shore = args[3]
        self.excel_file = '-'.join([arg[0], 'ACC', 'SHIFT', 'PLAN.xlsx')
        self.sheetName = '-'.join([self.team, self.year])
        self.format_excel_inst = formatExcel(self.excel_file, self.sheetName)

    """
    Save the excel dataframe to excel workbook.
    """
    def __save__(self, wb_inst):
        wb_inst.save(self.excel_file)

    """
    Fetch the existing dataframe contents fron the excel workbook.
    """
    def __existing_dataframe__(self):
        self.ex_df = pd.read_excel(self.excel_file, sheet_name=self.sheetName)
        return self.ex_df
    
    """
    Check if excel sheet exists, if not, creates a new excel sheet
    """
    def check_excel_exists(self):
        if not os.path.exists(self.excel_file):
            # If file not exists, create excel using pandas.
            self.log.info('Excel file not existing!!, Creating....')
            self.create_excel()
            self.log.info('File created....')
        return True

    """
    Create if excel file has the given sheet names.
    """
    def check_excel_sheet_exists(self, wb_inst):
        # Check if given sheet is present else create the sheet.
        if self.sheetName not in wb_inst.sheetnames:
            wb_inst.create_sheet(self.sheetName):
            # Save the excel file.
            self.__save__(wb_inst)
        # Just ensure excel has sheets (that you defined..)
        wb_existing = Workbook()
        self.log.debug('Listing out sheets present in the excel: ', self.excel_file)
        self.log.debug('Worksheets: ', wb_existing)
        for sheet in wb_existing:
            self.log.debug('Sheets Present in excel file are: ', sheet.title)

    """
    Create Excel if not exists.
    """
    def create_excel(self):
        initial_excel = Workbook()        
        if self.sheetName:
            self.log.debug('Adding user defined sheet name: ', str(self.sheetName))
            initial_excel.create_sheet(self.sheetName, 0)    # Insert at first position.
            self.log.debug('Remove default sheet name..')
            initial_excel.remove(initial_excel['Sheet'])
        initial_excel.save(self.excel_file)

    """
    Function to add overall shift plan headers for a month.
    """
    def add_headers_to_sheet(self):
        self.log.debug('MONTH : ' self.month)
        print(util.generate_month_days(self.month))
        if util.generate_month_days(self.month):
            days_list = util.generate_month_days(self.month)
            first_headers = ['Month', 'Shore', 'Day Number ->']
            first_headers.extend(days_list[0])
            first_headers.append('SHIFT DAYS STATS')
            pre_req_excel = pd.ExcelWriter(self.excel_file, engine='openpyxl')
            # We create first header here for the excel. (Month, Shore, Day Number -->, series of day no.s)
            # We create second header here for the excel. (Team Member, series of Week days)
            header_df1 = self.create_dataframe_for_headers(first_headers, pre_req_excel)
            second_header = util.generate_weekdays_for_month(self.year, self.month, len(days_list[0]))
            header_df2 = self.create_dataframe_for_headers(second_header, pre_req_excel)

            # Create a Pandas Excel writer using XlsxWriter as the engine.
            excel_writer = pd.ExcelWriter(self.excel_file, engine='openpyxl')
            # OPen existng workbook
            excel_writer.book = load_workbook(self.excel_file)
            # copy the existing contents in sheet
            excel_writer.sheets = dict((ex_content.title, ex_content) for ex_content in excel_writer.book.worksheets)
            # Read the excel file
            existing_df = self.__existing_dataframe__()
            
            if existing_df.empty:    # Means, checks if excel is completely empty.
                self.log.debug('empty excel!!')
                row_index = 0   # set starting row index to 0.
            
            elif existing_df.columns.values.all() and self.month:
                row_index = len(existing_df)+3   # set starting row index to 0.
            # Fill the headers to excel (2 rows are filled)
            header_df1.to_excel(excel_writer, sheet_name=self.sheetName, index=False, header=True, startrow=row_index)
            header_df2.to_excel(excel_writer, sheet_name=self.sheetName, index=False, header=True, startrow=row_index+1, startcol=2)            
            excel_writer.save()
            # Now perform excel formatting on saved headers.
            self.do_formatting_to_headers(days_list[0], header_df1)

    """
    Formats the Excel Headers with merging, bordering and color highlighting.
    """
    def do_formatting_to_headers(self, month_days, h_df):
        self.format_excel_inst.merge_required_header_cell(column_size=len(h_df.columns))
        _new_ex_df = self.__existing_dataframe__()
        # Evaluate month days last index.
        _month_end_col_index = const.THIRD_COLUMN + len(util.generate_month_days(self.month)[0])
        # Number of  shift category sum.
        _shift_sum_cols = len(cfg.get('shift_category')) # Adding 2 columns (Night & Evening Shifts)
        # Start index for shift category sum.
        _start_index_shift_sum = _month_end_col_index + 1
        # Evaluate shift category sum's last index.
        _shift_sum_end_index = _start_index_shift_sum + _shift_sum_cols
        # Start index for shift category hours.
        _start_index_shift_hours = _shift_sum_end_index + 1
        _indices_switcher = (
            31: _shift_sum_end_index + 1,
            30: _shift_sum_end_index,
            29: _shift_sum_end_index - 1,
            28: _shift_sum_end_index - 2
        )
        # Evaluate Shift Category Hours SUm Last Index.
        _end_index_shift_hours = _start_index_shift_hours + _shift_sum_cols
        self.format_excel_inst.set_background_color_multicells(srow=len(_new_ex_df), erow=len(_new_ex_df)+1, scol=const.FIRST_COLUMN \
            , ecol=const.SECOND_COLUMN, BGColor=const.HCOL1_TO_HCOL2)
        self.format_excel_inst.set_background_color_multicells(srow=len(_new_ex_df), erow=len(_new_ex_df)+1, scol=const.THIRD_COLUMN \
            , ecol=const.THIRD_COLUMN, BGColor=const.HCOL3)
        self.format_excel_inst.set_background_color_multicells(srow=len(_new_ex_df), erow=len(_new_ex_df)+1, scol=const.FOURTH_COLUMN \
        , ecol=_month_end_col_index, BGColor=const.HMONTH_DAYS)
        self.format_excel_inst.set_background_color_multicells(srow=len(_new_ex_df), erow=len(_new_ex_df), scol=_start_index_shift_sum, \
            ecol=_shift_sum_end_index, BGColor=const.HSHIFT_SUM)
        self.format_excel_inst.set_background_color_multicells(srow=len(_new_ex_df)+1, erow=len(_new_ex_df)+1, scol=_start_index_shift_sum, \
            ecol=_shift_sum_end_index, BGColor=const.HSHIFT_SUM_PARAMS)
        self.format_excel_inst.set_background_color_multicells(srow=len(_new_ex_df)+1, erow=len(_new_ex_df)+1, scol=_indices_switcher[len(month_days)], \)
            ecol=_end_index_shift_hours, BGColor=const.HSHIFT_SUM_PARAMS)

    """
    Staticmethod (No Relationship With Class): Create Dataframe for the given column headers.
    """
    def create_dataframe_for_headers(self, headers, excel_writer):
        days_keys_dict = dict.fromkeys(headers, None)
        # Create dataframe object for headers.
        dump_pre_reqs_df = pd.DataFrame(columns=headers)
        return dump_pre_reqs_df

    """
    Function to fill respective associate shift to respective cells (based on day number).
    """
    def fill_shift_row_cells(self, shift_df=None, shift_type='LEAVE', asso_name='Sabarish AC', start_row=2):
        __pos_index = cfg['shift_category'].index(shift_type) + 1
        if shift_type!='LEAVE' and shift_type!='HOLIDAY':
            self.log.debug(f'Positional index for {shift_type} is {__pos_index}')
            self.format_excel_inst.fill_cell_values(col=3, row=start_row, value=asso_name)
            for key, val in shift_df.items():
                self.format_excel_inst.fill_cell_values(col=(const.column_start + key), row=start_row, \
                    value=val, shift=shift_type)
        self.fill_shift_sum_row_cells(shift_df, shift_type, asso_name, start_row=start_row, ps_index=__pos_index)

    """
    Function to fill respective associate shift sum to respective cell.
    """
    def fill_shift_sum_row_cells(self, sh_df, sh_type, name, start_row=2, ps_index=0, shift_val=0):
        _ex_df = self.__existing_dataframe__()
        # Initialize _is_df flag variable to True, If no specific shift exists for an associate -> set it to False.
        _is_df = True
        days_in_month = util.generate_month_days(self.month)
        if not sh_df and sh_type!='LEAVE' and sh_type!='HOLIDAY':
            _is_df = False  # Setting flag to False as there are no shift plan existing.
            self.log.critical(f'No {sh_type} plan given for associate id: {str(name)}')
        # Fetch the counts of each shifts (ACC, GEN, OFF, LEAVE)
        _shift_count = self.fetch_shift_counts(_ex_df, _is_df)
        col_index = const.column_start + len(days_in_month[0]) + ps_index
        self.format_excel_inst.fill_cell_values(col=col_index, row=start_row, value=_shift_count[sh_type], shift=sh_type)
        self.log.debug(f'Filled Cells with {sh_type} to excel!!')


    """
    Convert user shift plan to object dictionary.
    """
    def convert_to_shift_dict(self, shift_plan='', gen_shift_keys=[], gen_shift_values=[]):
        _ex_df = self.__existing_dataframe__()
        
        if not gen_shift_keys and not gen_shift_values:
            shift_k = [int(shft.split(':')[0]) for idx, shft in enumerate(shift_plan.split(' ')) if shft.split(':')[0].isdigit()]
            shift_v = [shft for shft in shift_plan if shft!=' ' and shft!=':' and not shft.isdigit()]            
            shift_dict = dict(zip(shift_k, shift_v))
        elif gen_shift_keys and gen_shift_values:
            shift_dict = dict(zip(gen_shift_keys, gen_shift_values))
        else:
            shift_dict = {}
        return shift_dict

    """
    Append the associate shift plan to excel workbook (Avoid overwritting)
    """
    def append_contents_to_excel(self, asso_id, acc_plan, off_plan, leave_plan, tcs_holiday_plan):
        __asso_name = self.assoicates[asso_id]
        self.log.info(f'Appending contents for associate Name: {str(__asso_name)}, ID: {str(asso_id)}')
        _ex_df = self.__existing_dataframe__()
        #_acc = self.convert_to_shift_dict(shift_plan=self.accPlan[asso_id])
        self.log.debug(f'Filling ACC Plan for associate id: {str(__asso_name)}, ID: {str(asso_id)}, ACC Plan: {acc_plan}')
        # Fill the cells with corresponding associates shifts.
        self.fill_shift_row_cells(shift_df=acc_plan, shift_type=cfg.get('shift_category')[0], asso_name=self.assoicates[asso_id], start_row=len(_ex_df)+2)
        #_off = self.convert_to_shift_dict(self.offPlan[asso_id])
        # Fill OFF Plans
        self.log.debug(f'Filling OFF Plan for associate Name: {str(__asso_name)}, id: {str(asso_id)}, OFF Plan: {off_plan}')
        self.fill_shift_row_cells(shift_df=off_plan, shift_type=cfg.get('shift_category')[2], asso_name=self.assoicates[asso_id], start_row=len(_ex_df)+2)
        # Fill LEAVE Plans
        self.log.debug(f'Filling LEAVE Plan for associate Name: {str(__asso_name)}, id: {str(asso_id)}, LEAVE Plan: {leave_plan}')
        self.fill_shift_row_cells(shift_df=leave_plan, shift_type=cfg.get('shift_category')[3], asso_name=self.assoicates[asso_id], start_row=len(_ex_df)+2)
        # Fill HOLIDAY Plans
        self.log.debug(f'Filling HOLIDAY Plan for associate Name: {str(__asso_name)}, id: {str(asso_id)}, HOLIDAY Plan: {tcs_holiday_plan}')
        self.fill_shift_row_cells(shift_df=tcs_holiday_plan, shift_type=cfg.get('shift_category')[3], asso_name=self.assoicates[asso_id], start_row=len(_ex_df)+2)
        # Collect the days that has ACC, OFF, LEAVE & TCS_HOLIDAY shifts.
        _AOLHList = list(acc_plan.keys()) + list(off_plan.keys() + list(leave_plan.keys()) + list(holiday_plan.keys()))
        # Collect the days in a month.
        days_in_month = util.generate_month_days(self.month)
        # Find the difference b/w these 2 lists to collect remaining days as 'General' shifts.
        _gen_k = list(set(days_in_month[0]).difference(_AOLHList))
        _gen_v = ['G' for _k in _gen_k]
        _gen = self.convert_to_shift_dict(gen_shift_keys=_gen_k, gen_shift_values=_gen_v)
        self.log.debug(f'Filling GENERAL Plan for associate Name: {str(__asso_name)}, id: {str(asso_id)}, GENERAL Plan: {_gen}')
        self.fill_shift_row_cells(shift_df=_gen, shift_type=cfg.get('shift_category')[1], asso_name=self.assoicates[asso_id], start_row=len(_ex_df)+2)
        self.fill_shift_row_cells(shift_type=cfg.get('shift_category')[-2], asso_name=self.assoicates[asso_id], start_row=len(_ex_df)+2)
        self.fill_shift_row_cells(shift_type=cfg.get('shift_category')[-1], asso_name=self.assoicates[asso_id], start_row=len(_ex_df)+2)
        # Assign column index to fill sum of each shifts (default_col + no. of days in month + next cell)
        #_fill_shift_sum_column = const.column_start + len(days_in_month[0]) + 1

        if asso_id==len(self.assoicates)-1:
            self.format_excel_inst.merge_required_row_cells(asso_sum=len(self.assoicates), col=1, \
                value=self.month.upper(), month_days=len(days_in_month[0]))
            self.format_excel_inst.merge_required_row_cells(asso_sum=len(self.assoicates), col=2, \
                value=self.shore.upper(), month_days=len(days_in_month[0]))
        self.log.info(f'Associate Name: {str(__asso_name)}, ID: {str(asso_id)}: Appended Successfully!!') 
