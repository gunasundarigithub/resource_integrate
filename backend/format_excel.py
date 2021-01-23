"""
Copy Rights, All Rights Reserved 2020
File name: format_excel.py

Description: Script formats the excel workbook(fill cell values, cell alignments, borders, highlighting).

Code Changes:
Release Date    Revision Date   Changes By  Description
------------    -------------   ----------- ------------
                August 2020     Sabarish AC     Initial
                August 2020     Ganapathy R     Code Updation
"""
import os 
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font, colors
import sys
#sys.path.append('E:\\Saba Learning\\resource_integrate')
from util import util

# Setup Logging.
log = util.get_logger_obj()
conf = util.get_conf()
const = util.excelConstants()
class formatExcel():
    def __init__(self, excel_file, sheetName):
        self.excel = excel_file
        self.sheet = sheetName
        self.colorBG = util.excelConstants
        self.shift_category = ['ACC', 'NACC', 'EACC', 'GEN', 'OFF', 'LEAVE', 'HOLIDAY']
        self.shift_colorBG = [self.colorBG.ACC_COLOR, self.colorBG.NACC_COLOR, self.colorBG.EACC_COLOR, self.colorBG.GEN_COLOR, \
                self.colorBG.OFF_COLOR, self.colorBG.LEAVE_COLOR, self.colorBG.TCS_HOLIDAY_COLOR]
        self.header = False
        self.bold_f = False

    """
    Load Excel Workbook and return its object instance.
    """
    def __workbook__(self):
        wb = load_workbook(self.excel)
        return wb

    """
    Create Worksheet Object instance from work book Object with a given sheet name.
    """
    def __worksheet__ (self, wb_inst):
        ws =  wb_inst[self.sheet]
        return ws

    """
    Save the excel workbook changes.
    """   
    def __save__(self, wb_inst):
        wb_inst.save(self.excel)

    """
    Fetch existing data-frame contents from excel workbook.
    """
    def __existing_dataframe__(self):
        ex_df = pd.read_excel(self.excel, sheet_name=self.sheet, engine="openpyxl")
        return ex_df

    """
    Fetch GLOBAL FONT pattern based header flag.
    """
    def fetch_global_font(self, header, bold_f):
        GLOBAL_FONT  = Font(
                        name='calibri', 
                        size=11, 
                        bold=True if bold_f else False, 
                        italic=False, 
                        vertAlign=None, 
                        underline='none', 
                        strike=False, 
                        color=self.colorBG.HFONT if header else self.colorBG.VFONT) 
        return GLOBAL_FONT

    """
    Function to fill a cell with a given value.
    """
    def fill_cell_values(self, col=0, row=1, value='', shift='', bgcolor=''):
        _wb = self.__workbook__()
        _ws = self.__worksheet__(_wb)
        if not self.__existing_dataframe__().empty:
            # Fetch the current cell that you need to fill.
            print(f'value to e filled in the cell --> {value}')
            _curr_cell = _ws.cell(column=col, row=row, value=value)
            if col==1 or col==2:
                self.header = False    # Setting header flag to False, for Month, Shore and Associate name values.
                self.bold_f = True     # Setting bold flag to True, for Month, Shore and Associate name values
                _curr_cell.font = self.fetch_global_font(self.header, self.bold_f)
                _curr_cell.alignment = Alignment(horizontal='center', vertical='center')
                _curr_cell.fill = PatternFill(patternType="solid", fgColor=self.colorBG.MONTH_SHORE)
            else:
                self.header = False
                self.bold_f = False
                _curr_cell.font = self.fetch_global_font(self.header, self.bold_f)
                _curr_cell.alignment = Alignment(horizontal='center')
                if not shift:
                    _curr_cell.fill = PatternFill(patternType="solid", fgColor=self.colorBG.ASSO_COLOR)
                for _idx, _shift in enumerate(self.shift_category):
                    if str(value).upper()=='L':
                        _curr_cell.fill = PatternFill(patternType="solid", fgColor=self.colorBG.LEAVE_COLOR)
                    elif str(value).upper()=='H':
                        _curr_cell.fill = PatternFill(patternType="solid", fgColor=self.colorBG.TCS_HOLIDAY_COLOR)
                    elif shift == _shift:
                        _curr_cell.fill = PatternFill(patternType="solid", fgColor=self.shift_colorBG[_idx])
                    elif bgcolor:
                        _curr_cell.fill = PatternFill(patternType="solid", fgColor=bgcolor)
        
        # After formatting cell styings, Save it.
        self.__save__(_wb)
        # Now Set the borders to cell.
        self.set_borders_row(srow=row)

    """
    Function to add borders to rows/columns.
    """
    def set_borders_row(self, srow=1):
        _wb = self.__workbook__()
        _ws = self.__worksheet__(_wb)
        for _cell in _ws[srow:srow]:
            _cell.border = Border(
                top = Side(border_style='thin'),
                left = Side(border_style='thin'),
                right = Side(border_style='thin'),
                bottom = Side(border_style='thin')
            )
            self.__save__(_wb)

    """
    Set Borders to cells.
    """
    def set_borders_cell(self, srow=1, erow=1, col=1):
        _wb = self.__workbook__()
        _ws = self.__worksheet__(_wb)
        for _cell in _ws[srow, erow]:
            _cell[col] = Border(
                top = Side(border_style='thin'),
                left = Side(border_style='thin'),
                right = Side(border_style='thin'),
                bottom = Side(border_style='thin'),
            )
            self.__save__(_wb)

    """
    Function to fill back-ground color each cell based on associate shift. (A|N|E|G|O|L|H)
    """
    def set_background_color_multicells(self, srow=1, erow=1, scol=0, ecol=0, BGColor=''):
        _wb = self.__workbook__()
        _ws = self.__worksheet__(_wb)
        # If we are trying tofill color for 1 row. Make use of below functionality.
        if srow == erow:
            cell_range = _ws[srow:srow][scol-1:ecol]
            for _cell in cell_range:
                self.header = True
                self.bold_f = True
                _cell.font = self.fetch_global_font(self.header, self.bold_f)
                _cell.fill = PatternFill(patternType="solid", fgColor=BGColor)
        # If we are trying to fill color for multiple rows. Make use of below functionality.
        elif srow!=erow:
            # Multiple rows .. with cells
            cell_range = _ws[srow:erow]
            for _index, _cell_tuple in enumerate(cell_range):
                _cell = _cell_tuple[scol-1:ecol] 
                for _c in _cell:
                    self.header = True
                    self.bold_f = True
                    _c.font = self.fetch_global_font(self.header, self.bold_f) 
                    _c.fill = PatternFill(patternType="solid", fgColor=BGColor)
        self.__save__(_wb) 

    """
    Merge required header's cell
    """
    def merge_required_header_cell(self, column_size=0):
        _wb = self.__workbook__()
        _ws = self.__worksheet__(_wb)
        _srow = len(self.__existing_dataframe__())
        log.debug('Merging Required Header Cell')
        if not self.__existing_dataframe__().empty:
            _ws.merge_cells(start_row=_srow, start_column=1, end_row=_srow+1, end_column=1)
            _ws.merge_cells(start_row=_srow, start_column=2, end_row=_srow+1, end_column=2)
            # Merge sum shift sum header cells
            _ws.merge_cells(start_row=_srow, start_column=column_size, end_row=_srow, end_column=column_size+11)
        log.debug('Merged Done for Header Cell..')
        self.__save__(_wb)

    """
    Merge required  row's cell.
    """
    def merge_required_row_cells(self, asso_sum=3, col=0, value='', month_days=0):
        _wb = self.__workbook__()
        _ws = self.__worksheet__(_wb)  
        # Start row to be merged: (rows occupied in excel - no of associates in team) + 1
        # Note: In Excel, row index starts from 1, 2 ..... n
        _srow = len(self.__existing_dataframe__()) + 1 - (asso_sum) + 1
        _erow = len(self.__existing_dataframe__()) + 1
        log.debug(f'Merging required row cells')
        if not self.__existing_dataframe__().empty:
            _ws.merge_cells(start_row=_srow, start_column=col, end_row=_erow, end_column=col)
        self.__save__(_wb)
        log.debug(f'Merged required row cells')    
        self.fill_cell_values(col=col, row=_srow, value=value)
        for _r in range(_srow, _erow+1):
            self.set_borders_row(srow=_r)   