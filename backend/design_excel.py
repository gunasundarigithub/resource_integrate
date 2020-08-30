"""
Filename : design_excel.py

Description: Formating Excel  File

Modified by: Ganapathy R
"""

from openpyxl.styles import Color,Alignment, Border, Side, PatternFill, Font, colors 
from openpyxl import load_workbook
import pandas as pd
import openpyxlormat

"""
Format Excel headers.
"""
excel_file = 'TEST.xlsx'
import os

if not os.path.exists(excel_file):
    wb = Worknook()
    ws = Wb.active
    ws.title = 'CTPT'
    wb.save = (filename=excel_file)
    
Wb = load_workbook(excel_file)
excel_writer = pd.ExcelWriter(excel_file,engine='openpyxl')
# open existing workbook
excel_writer.book = wb
# copy existing contents in sheet
excel_writer.sheets = dict((ex_content.tittle, ex-content) for ex_content  in excel_writer.book.worksheets)
# Read the excel file
reader = pd.read(excel_file)
print('contents in excel: ', reader)
header = False
data -[['Sabarish AC', 'CTPT-CSE']]
df1 = pd.DataFrame([], Columns=['Name', 'Team', 'Company'])
df2 = pd.DataFrame([], Columns=['Stream', 'Comp'])

wb = workbook()
ws = wb.active

reader = pd.read_excel(excel_file), sheet_name='CTPT')
if reader.empty:
    for in dataframe_to_rows(df1, Index=False, header=True):
        ws.append(r)
        ws.tittle='CTPT'
        wb.save(excel_file)
        sheet= wb.worksheets[0]
        #ws = wb.active
    df2.to_excel(excel_writer, sheet_name='CTPT', index=False,header=False, startrow=2, startcol=2)
    excel_writer.save()
    GLOBAL_TTLE_FONT  = Font(name='Calibri', size=12, bold=True, italic=False, vertALIGN=None, underline='none', strike=False,  color='e64117') 
    wb.active.merge_cells('A1:A2')
    # Format rows
    for cell in ws['1:1']:
        cell.font = GLOBAL_TTLE_FONT
        cell.fill = PatternFill(start_color="ceffc7", end_color= "ceffc7",fill_type = "solid")
        row_count = sheet.max_row
        column_count = sheet.max_column
        print(row_count)
    wb.save(excel_file)    
    
        








    